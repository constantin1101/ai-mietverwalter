"""Export router — Excel multi-sheet workbook."""
from __future__ import annotations

import asyncio
import io
import logging
from datetime import date

from fastapi import APIRouter
from fastapi.responses import StreamingResponse

from app.core.dependencies import CurrentUser, SupabaseClient

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/export", tags=["export"])


def _run(db, fn):
    return asyncio.get_event_loop().run_in_executor(None, fn)


@router.get("/excel")
async def export_excel(
    current_user: CurrentUser,
    db: SupabaseClient,
) -> StreamingResponse:
    """Generate a multi-sheet Excel workbook for the user's portfolio."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl required. Run: uv add openpyxl") from exc

    uid = current_user.sub

    # Fetch all data
    units_resp = await _run(db, lambda: db.table("units")
        .select("id, property_id, unit_number, area_sqm, rooms, has_cellar, has_parking")
        .eq("user_id", uid).execute())
    units = units_resp.data or []

    prop_ids = list({u["property_id"] for u in units})
    props_resp = await _run(db, lambda: db.table("properties")
        .select("id, street, house_number, city, postal_code")
        .in_("id", prop_ids).execute()) if prop_ids else type("R", (), {"data": []})()
    props = {p["id"]: p for p in (props_resp.data or [])}

    leases_resp = await _run(db, lambda: db.table("leases")
        .select("*").eq("user_id", uid).eq("is_active", True).execute())
    leases = leases_resp.data or []
    lease_by_unit = {l["unit_id"]: l for l in leases}

    lease_ids = [l["id"] for l in leases]
    if lease_ids:
        lt_resp = await _run(db, lambda: db.table("lease_tenants")
            .select("lease_id, tenant_id, is_primary")
            .in_("lease_id", lease_ids).execute())
        lt_data = lt_resp.data or []
        primary_by_lease = {lt["lease_id"]: lt["tenant_id"] for lt in lt_data if lt["is_primary"]}
        all_tenant_ids = list({lt["tenant_id"] for lt in lt_data})
        t_resp = await _run(db, lambda: db.table("tenants")
            .select("*").in_("id", all_tenant_ids).execute())
        tenants_by_id = {t["id"]: t for t in (t_resp.data or [])}
    else:
        primary_by_lease = {}
        tenants_by_id = {}

    dl_resp = await _run(db, lambda: db.table("deadlines")
        .select("*").eq("user_id", uid).eq("is_completed", False)
        .order("due_date").execute())
    deadlines = dl_resp.data or []

    # Build workbook
    wb = Workbook()

    GREEN = "16A34A"
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill("solid", fgColor=GREEN)
    HEADER_ALIGN = Alignment(horizontal="left", vertical="center")

    def style_header(ws, headers: list[str]):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
        ws.row_dimensions[1].height = 22

    def autofit(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 40)

    # Sheet 1: Übersicht
    ws1 = wb.active
    ws1.title = "Übersicht"
    style_header(ws1, ["Adresse", "Einheit", "Fläche (m²)", "Zimmer", "Mieter", "Kaltmiete (€)", "NK (€)", "Mietbeginn", "Mietart"])
    for unit in units:
        lease = lease_by_unit.get(unit["id"])
        if not lease:
            continue
        prop = props.get(unit["property_id"], {})
        tid = primary_by_lease.get(lease["id"])
        tenant = tenants_by_id.get(tid, {}) if tid else {}
        ws1.append([
            f"{prop.get('street', '')} {prop.get('house_number', '')}, {prop.get('postal_code', '')} {prop.get('city', '')}",
            unit.get("unit_number") or "",
            unit.get("area_sqm"),
            unit.get("rooms"),
            f"{tenant.get('first_name', '')} {tenant.get('last_name', '')}".strip(),
            float(lease["base_rent"] or 0),
            float(lease["operating_costs"]) if lease.get("operating_costs") else "",
            str(lease.get("start_date", "")),
            lease.get("rent_type", "fixed"),
        ])
    autofit(ws1)

    # Sheet 2: Mieter
    ws2 = wb.create_sheet("Mieter")
    style_header(ws2, ["Vorname", "Nachname", "E-Mail", "Telefon", "Wohnung", "Mietbeginn"])
    for lease in leases:
        tid = primary_by_lease.get(lease["id"])
        tenant = tenants_by_id.get(tid, {}) if tid else {}
        unit = next((u for u in units if u["id"] == lease.get("unit_id")), {})
        prop = props.get(unit.get("property_id", ""), {})
        ws2.append([
            tenant.get("first_name", ""),
            tenant.get("last_name", ""),
            tenant.get("email", ""),
            tenant.get("phone", ""),
            f"{prop.get('street', '')} {prop.get('house_number', '')}" + (f" Nr. {unit.get('unit_number')}" if unit.get("unit_number") else ""),
            str(lease.get("start_date", "")),
        ])
    autofit(ws2)

    # Sheet 3: Finanzen
    ws3 = wb.create_sheet("Finanzen")
    style_header(ws3, ["Adresse", "Kaltmiete (€)", "Betriebskosten (€)", "Warmmiete (€)", "Kaution (€)", "Zahltag", "Zahlungsweise"])
    total_kalt = 0.0
    for lease in leases:
        unit = next((u for u in units if u["id"] == lease.get("unit_id")), {})
        prop = props.get(unit.get("property_id", ""), {})
        kalt = float(lease["base_rent"] or 0)
        nk = float(lease["operating_costs"]) if lease.get("operating_costs") else 0
        total_kalt += kalt
        ws3.append([
            f"{prop.get('street', '')} {prop.get('house_number', '')}",
            kalt, nk if nk else "", kalt + nk if nk else "",
            float(lease["deposit"]) if lease.get("deposit") else "",
            lease.get("payment_day", 1),
            "Überweisung" if lease.get("payment_method") == "transfer" else lease.get("payment_method", ""),
        ])
    ws3.append(["GESAMT", total_kalt, "", "", "", "", ""])
    ws3[f"A{ws3.max_row}"].font = Font(bold=True)
    ws3[f"B{ws3.max_row}"].font = Font(bold=True)
    autofit(ws3)

    # Sheet 4: Fristen
    ws4 = wb.create_sheet("Fristen")
    style_header(ws4, ["Titel", "Fällig am", "Typ", "Einheit"])
    for dl in deadlines:
        unit = next((u for u in units if u["id"] == dl.get("unit_id")), {})
        prop = props.get(unit.get("property_id", ""), {})
        ws4.append([
            dl.get("title", ""),
            str(dl.get("due_date", "")),
            dl.get("deadline_type", ""),
            f"{prop.get('street', '')} {prop.get('house_number', '')}".strip() or "—",
        ])
    autofit(ws4)

    # Stream
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"heimio_portfolio_{date.today().isoformat()}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

"""
Excel workbook builder for the Heimio portfolio export.
Produces an 8-sheet .xlsx file suitable for landlords and tax advisors.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ── Colour palette ─────────────────────────────────────────────────────────────
C_DARK    = "1C2B3A"   # header background
C_WHITE   = "FFFFFF"
C_GREEN   = "16A34A"   # primary accent
C_LIGHT   = "F9F9F7"   # alternating row
C_RED     = "FEE2E2"   # overdue
C_ORANGE  = "FEF3C7"   # ≤14 days
C_YELLOW  = "FEFCE8"   # ≤30 days
C_BLUE    = "EFF6FF"   # future / info
C_AMBER   = "FFFBEB"   # upcoming
C_BORDER  = "E5E7EB"   # thin border colour

# ── Number formats ─────────────────────────────────────────────────────────────
FMT_EURO = '#,##0.00 "€"'
FMT_DATE = 'DD.MM.YYYY'
FMT_PCT  = '+0.00%;-0.00%;0.00%'

# ── Tab colours per sheet ──────────────────────────────────────────────────────
TAB_COLORS = [
    C_DARK, C_GREEN, "0369A1", "7C3AED",
    "0891B2", "DC2626", "D97706", "6B7280",
]

# ── Reusable style helpers ─────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = "000000", size: int = 11) -> Font:
    return Font(bold=bold, color=color, size=size)


def _border_thin() -> Border:
    side = Side(style="thin", color=C_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


def _style_header_row(ws, col_count: int, row: int = 1) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font   = _font(bold=True, color=C_WHITE, size=10)
        cell.fill   = _fill(C_DARK)
        cell.alignment = _left()
        cell.border = _border_thin()
    ws.row_dimensions[row].height = 22


def _style_data_row(ws, row: int, col_count: int, alt: bool, fill_override: str | None = None) -> None:
    bg = fill_override or (C_LIGHT if alt else C_WHITE)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        if fill_override or alt:
            cell.fill = _fill(bg)
        cell.border = _border_thin()
        cell.alignment = _left()


def _freeze(ws, after_row: int = 1) -> None:
    ws.freeze_panes = ws.cell(row=after_row + 1, column=1)


def _autofit(ws, max_width: int = 45) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, max_width)


def _tab(ws, idx: int) -> None:
    ws.sheet_properties.tabColor = TAB_COLORS[idx % len(TAB_COLORS)]


def _fmt_date(val: Any) -> str:
    if not val:
        return ""
    try:
        if isinstance(val, (date, datetime)):
            return val.strftime("%d.%m.%Y")
        return datetime.fromisoformat(str(val)).strftime("%d.%m.%Y")
    except (ValueError, TypeError):
        return str(val)


def _fmt_rent_type(t: str | None) -> str:
    return {"fixed": "Festmiete", "indexed": "Indexmiete", "graduated": "Staffelmiete"}.get(t or "", t or "")


def _fmt_payment(m: str | None) -> str:
    return {"transfer": "Banküberweisung", "direct_debit": "SEPA-Lastschrift"}.get(m or "", m or "")


def _fmt_doc_type(t: str | None) -> str:
    return {
        "lease_contract": "Mietvertrag", "handover_protocol": "Übergabeprotokoll",
        "energy_certificate": "Energieausweis", "insurance_policy": "Versicherung",
        "invoice": "Rechnung", "rent_increase": "Mieterhöhung",
        "utility_statement": "NK-Abrechnung", "correspondence": "Korrespondenz",
        "other": "Sonstiges",
    }.get(t or "", t or "Sonstiges")


def _fmt_deadline_type(t: str | None) -> str:
    return {
        "rent_adjustment": "Mietanpassung", "lease_termination": "Vertragsende",
        "notice_period": "Kündigungsfrist", "inspection": "Besichtigung",
        "insurance_renewal": "Versicherung", "utility_statement": "NK-Abrechnung",
        "tax_deadline": "Steuertermin", "custom": "Individuell",
    }.get(t or "", t or "")


def _days_until(due_date_str: str) -> int:
    try:
        due = date.fromisoformat(str(due_date_str))
        return (due - date.today()).days
    except (ValueError, TypeError):
        return 9999


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _sheet_deckblatt(wb: Workbook, data: dict, export_date: date) -> None:
    ws = wb.active
    ws.title = "Deckblatt"
    _tab(ws, 0)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28

    def title_row(row: int, text: str, size: int = 11, bold: bool = False, color: str = "000000") -> None:
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = _font(bold=bold, color=color, size=size)
        cell.alignment = _left()

    def kv_row(row: int, key: str, value: Any, value_color: str = "000000") -> None:
        k = ws.cell(row=row, column=1, value=key)
        k.font = _font(bold=False, color="6B7280", size=10)
        v = ws.cell(row=row, column=2, value=value)
        v.font = _font(bold=True, color=value_color, size=10)

    # Title block
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 36
    title = ws.cell(row=2, column=1, value="Heimio")
    title.font = Font(bold=True, color=C_GREEN, size=26)
    sub = ws.cell(row=3, column=1, value="Portfolio-Export")
    sub.font = _font(bold=False, color=C_DARK, size=14)

    ws.row_dimensions[5].height = 10
    title_row(6, "Export-Informationen", bold=True, size=11, color=C_DARK)
    kv_row(7, "Erstellt am",      export_date.strftime("%d.%m.%Y"))
    kv_row(8, "Einheiten gesamt", data["kpi_units"])
    kv_row(9, "Gesamtkaltmiete",  f"{data['kpi_total_rent']:,.2f} €".replace(",", "X").replace(".", ",").replace("X", "."), C_GREEN)
    kv_row(10, "Gesamtwarmmiete", f"{data['kpi_total_warm']:,.2f} €".replace(",", "X").replace(".", ",").replace("X", "."))
    kv_row(11, "Offene Fristen",  data["kpi_deadlines"])

    ws.row_dimensions[13].height = 10
    title_row(14, "Enthaltene Sheets", bold=True, size=11, color=C_DARK)
    sheets_info = [
        ("2 · Portfolio-Übersicht",  "Alle Einheiten mit Miete, Mieter, Vertragsdaten"),
        ("3 · Mietverträge",         "Alle Vertragsdetails inkl. Sondervereinbarungen"),
        ("4 · Mieter",               "Kontaktdaten aller Mieter"),
        ("5 · Finanzen",             "Kalt-/Warmmiete, Kaution, €/m², Jahressummen"),
        ("6 · Fristen & Termine",    "Alle offenen Fristen, farblich nach Dringlichkeit"),
        ("7 · Mietentwicklung",      "Staffel- und Indexmiete-Verlauf + Prognose"),
        ("8 · Dokumente",            "Dokumenten-Inventar aller Einheiten"),
    ]
    for i, (name, desc) in enumerate(sheets_info):
        row = 15 + i
        ws.cell(row=row, column=1, value=name).font = _font(bold=True, size=10)
        ws.cell(row=row, column=2, value=desc).font = _font(color="6B7280", size=10)

    ws.row_dimensions[23].height = 10
    footer = ws.cell(row=24, column=1, value="Erstellt mit Heimio · heimio.de · Alle Angaben ohne Gewähr")
    footer.font = _font(color="9CA3AF", size=9)


def _sheet_portfolio(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Portfolio-Übersicht")
    _tab(ws, 1)

    headers = [
        "Adresse", "Einheit", "PLZ", "Stadt",
        "Mieter", "Kaltmiete (€)", "NK VP (€)", "Warmmiete (€)",
        "€/m²", "Zimmer", "Fläche (m²)", "Mietart",
        "Mietbeginn", "Vertragsende", "Kaution (€)",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    _freeze(ws)

    total_kalt = total_warm = 0.0

    for i, unit in enumerate(data["units"], start=1):
        lease = data["lease_by_unit"].get(unit["id"])
        if not lease:
            continue
        prop  = data["props"].get(unit.get("property_id", ""), {})
        tid   = data["primary_by_lease"].get(lease["id"])
        tenant = data["tenants_by_id"].get(tid, {}) if tid else {}

        kalt  = float(lease.get("base_rent") or 0)
        nk    = float(lease.get("operating_costs") or 0)
        warm  = kalt + nk
        area  = float(unit.get("area_sqm") or 0)
        per_m2 = round(kalt / area, 2) if area else None
        total_kalt += kalt
        total_warm += warm

        row_data = [
            f"{prop.get('street', '')} {prop.get('house_number', '')}".strip(),
            unit.get("unit_number") or "",
            prop.get("postal_code", ""),
            prop.get("city", ""),
            f"{tenant.get('first_name', '')} {tenant.get('last_name', '')}".strip(),
            kalt, nk or None, warm,
            per_m2,
            unit.get("rooms"),
            area or None,
            _fmt_rent_type(lease.get("rent_type")),
            _fmt_date(lease.get("start_date")),
            _fmt_date(lease.get("end_date")),
            float(lease.get("deposit") or 0) or None,
        ]
        ws.append(row_data)
        alt = i % 2 == 0
        _style_data_row(ws, ws.max_row, len(headers), alt)

        for col in [6, 7, 8, 9, 15]:
            ws.cell(row=ws.max_row, column=col).number_format = FMT_EURO

    # Totals row
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="GESAMT").font = _font(bold=True)
    ws.cell(row=total_row, column=6, value=total_kalt).number_format = FMT_EURO
    ws.cell(row=total_row, column=8, value=total_warm).number_format = FMT_EURO
    for col in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=col).font = _font(bold=True)
        ws.cell(row=total_row, column=col).fill = _fill("F0FDF4")
        ws.cell(row=total_row, column=col).border = _border_thin()

    _autofit(ws)


def _sheet_mietvertraege(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Mietverträge")
    _tab(ws, 2)

    headers = [
        "Adresse", "Einheit", "Mieter",
        "Mietbeginn", "Vertragsende", "Befristet",
        "Kündigungsfrist (Mo.)", "Zahltag", "Zahlungsweise",
        "Mietart", "Haustiere", "Untermiete",
        "Schönheitsreparaturen", "Kaution (€)",
        "Index-Typ", "Index-Basiswert", "Index-Basisdatum", "Index-Intervall (Mo.)",
        "Extraktions-Konfidenz",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    _freeze(ws)

    for i, lease in enumerate(data["leases"], start=1):
        unit  = next((u for u in data["units"] if u["id"] == lease.get("unit_id")), {})
        prop  = data["props"].get(unit.get("property_id", ""), {})
        tid   = data["primary_by_lease"].get(lease["id"])
        tenant = data["tenants_by_id"].get(tid, {}) if tid else {}

        row_data = [
            f"{prop.get('street', '')} {prop.get('house_number', '')}".strip(),
            unit.get("unit_number") or "",
            f"{tenant.get('first_name', '')} {tenant.get('last_name', '')}".strip(),
            _fmt_date(lease.get("start_date")),
            _fmt_date(lease.get("end_date")),
            "Ja" if lease.get("is_fixed_term") else "Nein",
            lease.get("notice_period_months", 3),
            lease.get("payment_day", 1),
            _fmt_payment(lease.get("payment_method")),
            _fmt_rent_type(lease.get("rent_type")),
            "Erlaubt" if lease.get("pets_allowed") is True else ("Nicht erlaubt" if lease.get("pets_allowed") is False else ""),
            "Erlaubt" if lease.get("subletting_allowed") is True else ("Nicht erlaubt" if lease.get("subletting_allowed") is False else ""),
            lease.get("cosmetic_repairs_clause") or "",
            float(lease.get("deposit") or 0) or None,
            lease.get("index_type") or "",
            float(lease.get("index_base_value") or 0) or None,
            _fmt_date(lease.get("index_base_date")),
            lease.get("index_adjustment_interval_months") or None,
            round(float(lease.get("extraction_confidence") or 0), 2) or None,
        ]
        ws.append(row_data)
        _style_data_row(ws, ws.max_row, len(headers), i % 2 == 0)
        ws.cell(row=ws.max_row, column=14).number_format = FMT_EURO

    _autofit(ws)


def _sheet_mieter(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Mieter")
    _tab(ws, 3)

    headers = ["Vorname", "Nachname", "E-Mail", "Telefon", "Einheit / Adresse", "Rolle", "Mietbeginn", "Kaution (€)"]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    _freeze(ws)

    rows_written = 0
    for lt in data["all_lt"]:
        tenant = data["tenants_by_id"].get(lt["tenant_id"], {})
        lease  = next((l for l in data["leases"] if l["id"] == lt["lease_id"]), {})
        unit   = next((u for u in data["units"] if u["id"] == lease.get("unit_id")), {})
        prop   = data["props"].get(unit.get("property_id", ""), {})

        addr = f"{prop.get('street', '')} {prop.get('house_number', '')}".strip()
        if unit.get("unit_number"):
            addr += f" · {unit['unit_number']}"

        rows_written += 1
        row_data = [
            tenant.get("first_name", ""),
            tenant.get("last_name", ""),
            tenant.get("email", ""),
            tenant.get("phone", ""),
            addr,
            "Hauptmieter" if lt.get("is_primary") else "Mitmieter",
            _fmt_date(lease.get("start_date")),
            float(lease.get("deposit") or 0) or None,
        ]
        ws.append(row_data)
        _style_data_row(ws, ws.max_row, len(headers), rows_written % 2 == 0)
        ws.cell(row=ws.max_row, column=8).number_format = FMT_EURO

    _autofit(ws)


def _sheet_finanzen(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Finanzen")
    _tab(ws, 4)

    headers = [
        "Adresse", "Mieter",
        "Kaltmiete / Mo. (€)", "NK VP / Mo. (€)", "Warmmiete / Mo. (€)",
        "Jahres-Kaltmiete (€)", "Jahres-Warmmiete (€)",
        "Kaution (€)", "Kaution (Monatskaltm.)", "€/m²",
        "Zimmer", "Fläche (m²)", "Zahltag", "Zahlungsweise",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    _freeze(ws)

    total_kalt = total_warm = 0.0

    for i, lease in enumerate(data["leases"], start=1):
        unit  = next((u for u in data["units"] if u["id"] == lease.get("unit_id")), {})
        prop  = data["props"].get(unit.get("property_id", ""), {})
        tid   = data["primary_by_lease"].get(lease["id"])
        tenant = data["tenants_by_id"].get(tid, {}) if tid else {}

        kalt  = float(lease.get("base_rent") or 0)
        nk    = float(lease.get("operating_costs") or 0)
        warm  = kalt + nk
        area  = float(unit.get("area_sqm") or 0)
        per_m2 = round(kalt / area, 2) if area else None
        kaution = float(lease.get("deposit") or 0)
        kaution_mo = round(kaution / kalt, 1) if kalt else None
        total_kalt += kalt
        total_warm += warm

        row_data = [
            f"{prop.get('street', '')} {prop.get('house_number', '')}".strip(),
            f"{tenant.get('first_name', '')} {tenant.get('last_name', '')}".strip(),
            kalt, nk or None, warm,
            kalt * 12, warm * 12,
            kaution or None, kaution_mo,
            per_m2,
            unit.get("rooms"),
            area or None,
            lease.get("payment_day", 1),
            _fmt_payment(lease.get("payment_method")),
        ]
        ws.append(row_data)
        _style_data_row(ws, ws.max_row, len(headers), i % 2 == 0)
        for col in [3, 4, 5, 6, 7, 8, 10]:
            ws.cell(row=ws.max_row, column=col).number_format = FMT_EURO

    # Totals
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="GESAMT").font = _font(bold=True)
    for col, val in [(3, total_kalt), (5, total_warm), (6, total_kalt * 12), (7, total_warm * 12)]:
        c = ws.cell(row=total_row, column=col, value=val)
        c.font = _font(bold=True)
        c.number_format = FMT_EURO
    for col in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=col).fill = _fill("F0FDF4")
        ws.cell(row=total_row, column=col).border = _border_thin()

    _autofit(ws)


def _sheet_fristen(wb: Workbook, data: dict, date_from: date | None, date_to: date | None) -> None:
    ws = wb.create_sheet("Fristen & Termine")
    _tab(ws, 5)

    headers = ["Titel", "Fällig am", "Tage bis Fälligkeit", "Typ", "Einheit", "Mieter", "Status"]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    _freeze(ws)

    deadlines = data["deadlines"]
    if date_from:
        deadlines = [d for d in deadlines if d.get("due_date") and str(d["due_date"]) >= date_from.isoformat()]
    if date_to:
        deadlines = [d for d in deadlines if d.get("due_date") and str(d["due_date"]) <= date_to.isoformat()]

    for i, dl in enumerate(deadlines, start=1):
        unit  = next((u for u in data["units"] if u["id"] == dl.get("unit_id")), {})
        prop  = data["props"].get(unit.get("property_id", ""), {})
        tid   = data["primary_by_lease"].get(data["lease_by_unit"].get(unit.get("id"), {}).get("id", ""))
        tenant = data["tenants_by_id"].get(tid, {}) if tid else {}

        days = _days_until(dl.get("due_date", ""))
        addr = f"{prop.get('street', '')} {prop.get('house_number', '')}".strip()

        row_data = [
            dl.get("title", ""),
            _fmt_date(dl.get("due_date")),
            days if days < 9999 else None,
            _fmt_deadline_type(dl.get("deadline_type")),
            addr or "—",
            f"{tenant.get('first_name', '')} {tenant.get('last_name', '')}".strip() or "—",
            "Überfällig" if days < 0 else "Offen",
        ]
        ws.append(row_data)

        # Color by urgency
        if days < 0:
            fill_color = C_RED
        elif days <= 14:
            fill_color = C_ORANGE
        elif days <= 30:
            fill_color = C_YELLOW
        else:
            fill_color = None

        _style_data_row(ws, ws.max_row, len(headers), i % 2 == 0, fill_override=fill_color)

    _autofit(ws)


def _sheet_mietentwicklung(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Mietentwicklung")
    _tab(ws, 6)

    headers = [
        "Einheit", "Mieter", "Typ",
        "Gültig ab", "Alte Miete (€)", "Neue Miete (€)",
        "Veränderung (€)", "Veränderung (%)", "Anschreiben verschickt",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    _freeze(ws)

    today = date.today()

    # Build current rent lookup by lease_id
    current_rent: dict[str, float] = {l["id"]: float(l.get("base_rent") or 0) for l in data["leases"]}
    lease_by_id: dict[str, dict] = {l["id"]: l for l in data["leases"]}

    # Sort adjustments by lease + effective_date
    adjustments = sorted(data["rent_adjustments"], key=lambda r: (r.get("lease_id", ""), r.get("effective_date", "")))

    # Track running rent per lease for delta calculation
    running: dict[str, float] = {}

    row_num = 0
    for adj in adjustments:
        lease  = lease_by_id.get(adj.get("lease_id", ""), {})
        unit   = next((u for u in data["units"] if u["id"] == lease.get("unit_id")), {})
        prop   = data["props"].get(unit.get("property_id", ""), {})
        tid    = data["primary_by_lease"].get(lease.get("id", ""))
        tenant = data["tenants_by_id"].get(tid, {}) if tid else {}

        lease_id  = adj.get("lease_id", "")
        new_rent  = float(adj.get("new_base_rent") or 0)
        old_rent  = running.get(lease_id, current_rent.get(lease_id, 0.0))
        running[lease_id] = new_rent
        delta     = new_rent - old_rent
        delta_pct = (delta / old_rent) if old_rent else 0

        eff_date = adj.get("effective_date", "")
        is_future = eff_date and str(eff_date) > today.isoformat()

        row_data = [
            f"{prop.get('street', '')} {prop.get('house_number', '')}".strip(),
            f"{tenant.get('first_name', '')} {tenant.get('last_name', '')}".strip(),
            {"graduated": "Staffelmiete", "index": "Indexmiete", "manual": "Manuell", "mietspiegel": "Mietspiegel"}.get(adj.get("adjustment_type", ""), adj.get("adjustment_type", "")),
            _fmt_date(eff_date),
            old_rent or None,
            new_rent,
            delta or None,
            delta_pct or None,
            "Ja" if adj.get("notice_sent") else "Nein",
        ]

        row_num += 1
        ws.append(row_data)
        _style_data_row(ws, ws.max_row, len(headers), row_num % 2 == 0,
                        fill_override=C_BLUE if is_future else None)

        for col in [5, 6, 7]:
            ws.cell(row=ws.max_row, column=col).number_format = FMT_EURO
        ws.cell(row=ws.max_row, column=8).number_format = FMT_PCT

    if row_num == 0:
        ws.cell(row=2, column=1, value="Keine Mietanpassungen vorhanden").font = _font(color="9CA3AF")

    _autofit(ws)


def _sheet_dokumente(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Dokumente")
    _tab(ws, 7)

    headers = ["Dateiname", "Typ", "Einheit", "Mieter", "Hochgeladen am", "Dateigröße", "Status"]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    _freeze(ws)

    for i, doc in enumerate(data["documents"], start=1):
        unit  = next((u for u in data["units"] if u["id"] == doc.get("unit_id")), {}) if doc.get("unit_id") else {}
        prop  = data["props"].get(unit.get("property_id", ""), {}) if unit else {}
        tid   = data["primary_by_lease"].get(data["lease_by_unit"].get(unit.get("id"), {}).get("id", "")) if unit else None
        tenant = data["tenants_by_id"].get(tid, {}) if tid else {}

        size_bytes = doc.get("file_size_bytes")
        if size_bytes:
            size_str = f"{size_bytes / (1024*1024):.1f} MB" if size_bytes >= 1024*1024 else f"{size_bytes / 1024:.0f} KB"
        else:
            size_str = ""

        status_map = {"uploaded": "Hochgeladen", "processing": "Wird verarbeitet", "extracted": "Verarbeitet", "error": "Fehler"}

        row_data = [
            doc.get("filename", ""),
            _fmt_doc_type(doc.get("document_type")),
            f"{prop.get('street', '')} {prop.get('house_number', '')}".strip() or "—",
            f"{tenant.get('first_name', '')} {tenant.get('last_name', '')}".strip() or "—",
            _fmt_date(doc.get("created_at")),
            size_str,
            status_map.get(doc.get("status", ""), doc.get("status", "")),
        ]
        ws.append(row_data)
        _style_data_row(ws, ws.max_row, len(headers), i % 2 == 0)

    _autofit(ws)


# ── Main entry point ───────────────────────────────────────────────────────────

def build_workbook(
    data: dict,
    date_from: date | None = None,
    date_to: date | None = None,
) -> bytes:
    """
    Build the complete 8-sheet Excel workbook and return the raw bytes.

    ``data`` must contain the keys populated by the router's fetch logic:
    units, props, leases, lease_by_unit, all_lt, primary_by_lease,
    tenants_by_id, deadlines, rent_adjustments, documents.
    """
    wb = Workbook()
    export_date = date.today()

    # Pre-compute KPIs for cover page
    total_kalt = sum(float(l.get("base_rent") or 0) for l in data["leases"])
    total_warm = total_kalt + sum(float(l.get("operating_costs") or 0) for l in data["leases"])
    data["kpi_units"]      = len(data["leases"])
    data["kpi_total_rent"] = total_kalt
    data["kpi_total_warm"] = total_warm
    data["kpi_deadlines"]  = len(data["deadlines"])

    _sheet_deckblatt(wb, data, export_date)
    _sheet_portfolio(wb, data)
    _sheet_mietvertraege(wb, data)
    _sheet_mieter(wb, data)
    _sheet_finanzen(wb, data)
    _sheet_fristen(wb, data, date_from, date_to)
    _sheet_mietentwicklung(wb, data)
    _sheet_dokumente(wb, data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
